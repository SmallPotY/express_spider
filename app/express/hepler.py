# coding=utf-8
from datetime import time
import openpyxl
from sqlalchemy import exists

from app.models import Express
from app import db


class Excel:

    def __init__(self, f):
        self.wb = openpyxl.load_workbook(f)
        self.sheet = self.wb.active
        self.max_column = self.sheet.max_column  # 最大列
        self.max_row = self.sheet.max_row  # 做大行
        self.row = 2

    def next(self):
        self.item = [i.value for i in self.sheet[self.row]]
        self.row += 1
        return self.item


if __name__ == '__main__':

    a = Excel('模板.xlsx')

    for i in range(a.max_row - 1):
        item = a.next()
        s = Express(express_order=item[0])

        if not db.session.query(exists().where(Express.express_order == item[0])).scalar():
            # print('记录存在')
            db.session.add(s)
            # print('插入成功')

    db.session.commit()
    # print('写入完成')
